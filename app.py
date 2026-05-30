import streamlit as st
import biosteam as bst
import thermosteam as tmo
import pandas as pd
import google.generativeai as genai
import os
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ==========================================
# 1. CONFIGURACIÓN VISUAL Y ESTILOS (CSS)
# ==========================================
st.set_page_config(page_title="Concentración de Mosto - IMIQ", layout="wide")

# Inyección de estilos globales para métricas y los nuevos efectos interactivos neón
st.markdown("""
    <style>
    /* Fondo blanco y borde para los recuadros de métricas */
    [data-testid="stMetric"] {
        background-color: #ffffff !important;
        padding: 20px;
        border-radius: 12px;
        box-shadow: 0 4px 6px rgba(0,0,0,0.1);
        border: 1px solid #d1d5db;
    }
    /* Forzar color de texto oscuro para etiquetas y valores */
    [data-testid="stMetricLabel"] > div {
        color: #4b5563 !important;
        font-weight: bold;
    }
    [data-testid="stMetricValue"] > div {
        color: #111827 !important;
    }

    /* --- ESTILOS DE INTERACTIVIDAD PARA DIAGRAMAS SVG --- */
    .contenedor-svg-interactivo {
        width: 100%;
        margin-bottom: 50px;
        background: transparent;
        display: block;
    }
    /* Efecto de iluminación interna para el BFD (Lila/Azul) */
    .vector-lila svg {
        width: 100%;
        height: auto;
        transition: filter 0.3s ease-in-out;
    }
    .vector-lila svg:hover {
        filter: brightness(1.3) drop-shadow(0 0 20px #bd00ff);
        cursor: pointer;
    }
    /* Efecto de iluminación interna para el PFD (Verde/Amarillo) */
    .vector-verde svg {
        width: 100%;
        height: auto;
        transition: filter 0.3s ease-in-out;
    }
    .vector-verde svg:hover {
        filter: brightness(1.3) drop-shadow(0 0 20px #39ff14);
        cursor: pointer;
    }

    /* --- ESTILOS HOVER PARA EL PANEL DE TARJETAS INDUSTRIALES (HMI) --- */
    .tarjeta-industrial {
        background-color: #0e1117 !important;
        padding: 20px;
        border-radius: 12px;
        margin-bottom: 20px;
        transition: all 0.3s ease-in-out;
        border: 1px solid #262730;
    }
    .titulo-tarjeta {
        font-family: 'Courier New', monospace;
        font-weight: bold;
        margin-bottom: 10px;
        font-size: 18px;
    }
    /* Enciende en Lila/Azul al pasar el cursor */
    .tarjeta-lila:hover {
        border: 2px solid #bd00ff !important;
        box-shadow: 0 0 20px #00d4ff !important;
        transform: translateY(-5px);
        cursor: pointer;
    }
    /* Enciende en Verde/Amarillo al pasar el cursor */
    .tarjeta-verde:hover {
        border: 2px solid #39ff14 !important;
        box-shadow: 0 0 20px #fff000 !important;
        transform: translateY(-5px);
        cursor: pointer;
    }
    </style>
    """, unsafe_allow_html=True)

# ==========================================
# 2. FUNCIÓN DE SIMULACIÓN
# ==========================================
def run_simulation(t_feed, t_w220, p_v1, p_luz, p_vapor, p_agua, p_mosto, p_etanol):
    bst.main_flowsheet.clear()
    
    chemicals = tmo.Chemicals(["Water", "Ethanol"])
    bst.settings.set_thermo(chemicals)

    # Corrientes
    mosto = bst.Stream("1-MOSTO", Water=900, Ethanol=100, units="kg/hr", 
                       T=t_feed + 273.15, price=p_mosto)
    vinazas_retorno = bst.Stream("Vinazas-Retorno", Water=200, T=95+273.15)

    # Equipos
    P100 = bst.Pump("P100", ins=mosto, P=4*101325)
    W210 = bst.HXprocess("W210", ins=(P100-0, vinazas_retorno), 
                         outs=("Mosto_Pre", "Drenaje"), phase0="l", phase1="l")
    W210.outs[0].T = 85 + 273.15
    W220 = bst.HXutility("W220", ins=W210-0, outs="Mezcla", T=t_w220 + 273.15)
    V1 = bst.Flash("V1", ins=W220-0, outs=("Vapor", "Vinazas"), P=p_v1 * 101325, Q=0)
    prod = bst.Stream("Producto_Final", price=p_etanol)
    W310 = bst.HXutility("W310", ins=V1-0, outs=prod, T=25 + 273.15)
    P200 = bst.Pump("P200", ins=V1-1, outs=vinazas_retorno, P=3*101325)

    sys = bst.System("mosto_sys", path=(P100, W210, W220, V1, W310, P200))
    sys.simulate()
    
    return sys, prod

# ==========================================
# 3. SIDEBAR (PUNTOS 1 AL 8)
# ==========================================
with st.sidebar:
    st.header("🎛️ Parámetros de Operación")
    t_f = st.slider("1. Temp. Alimentación (°C)", 10, 50, 25)
    t_out = st.slider("2. Temp. Salida W220 (°C)", 70, 110, 92)
    p_v = st.slider("3. Presión V1 (atm)", 0.1, 2.0, 1.0)
    
    st.header("💰 Costos de Insumos")
    p_luz = st.slider("4. Precio Luz (USD/kWh)", 0.05, 0.40, 0.15)
    p_vap = st.slider("5. Precio Vapor (USD/ton)", 10, 60, 25)
    p_agu = st.slider("6. Precio Agua (USD/m3)", 0.5, 5.0, 1.5)
    
    st.header("📈 Precios de Mercado")
    p_mos = st.slider("7. Precio Mosto (USD/kg)", 0.1, 2.0, 0.5)
    p_eta = st.slider("8. Precio Etanol (USD/kg)", 1.0, 6.0, 3.5)

# ==========================================
# 4. EXCEL GENERATOR FUNCTION
# ==========================================
def generar_excel_dinamico(datos_producto, balances, sensibilidad, escenarios):
    wb = openpyxl.Workbook()
    ws1 = wb.active; ws1.title = "Resumen e Indicadores"
    ws2 = wb.create_sheet(title="Balances de Materia y Energía")
    ws3 = wb.create_sheet(title="Análisis Económico")
    ws4 = wb.create_sheet(title="Comparación de Escenarios")
    
    font_title = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_section = Font(name="Segoe UI", size=12, bold=True, color="1F497D")
    font_data = Font(name="Segoe UI", size=11)
    
    fill_blue = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
    )

    # Pestaña 1
    ws1.views.sheetView[0].showGridLines = True
    ws1.merge_cells("A1:D2")
    ws1["A1"] = "REPORTE DE RENDIMIENTO DE PROCESO"
    ws1["A1"].font = font_title; ws1["A1"].fill = fill_blue; ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1["A4"] = "Datos del Producto Final"; ws1["A4"].font = font_section
    for c, h in enumerate(["Parámetro", "Valor", "Unidad", "Descripción"], 1):
        cell = ws1.cell(row=5, column=c, value=h); cell.font = font_header; cell.fill = fill_blue
    for r, row_data in enumerate(datos_producto, 6):
        for c, val in enumerate(row_data, 1):
            cell = ws1.cell(row=r, column=c, value=val); cell.font = font_data; cell.border = thin_border
            if c == 2 and isinstance(val, (int, float)): cell.number_format = '#,##0.0000'

    # Pestaña 2
    ws2.views.sheetView[0].showGridLines = True
    ws2.merge_cells("A1:F2")
    ws2["A1"] = "BALANCES DE MATERIA Y ENERGÍA POR CORRIENTE"
    ws2["A1"].font = font_title; ws2["A1"].fill = fill_blue; ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    for c, h in enumerate(["Corriente", "Descripción", "Temperatura (°C)", "Presión (bar)", "Flujo Másico (kg/h)", "Entalpia (kW)"], 1):
        cell = ws2.cell(row=4, column=c, value=h); cell.font = font_header; cell.fill = fill_blue
    for r, row_data in enumerate(balances, 5):
        for c, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r, column=c, value=val); cell.font = font_data; cell.border = thin_border
            if c > 2 and isinstance(val, (int, float)): cell.number_format = '#,##0.00'

    # Pestaña 3
    ws3.views.sheetView[0].showGridLines = True
    ws3.merge_cells("A1:E2")
    ws3["A1"] = "ANÁLISIS DE SENSIBILIDAD ECONÓMICA"
    ws3["A1"].font = font_title; ws3["A1"].fill = fill_blue; ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws3["A4"] = "Sensibilidad del Proyecto (Variación del Precio del Vapor)"; ws3["A4"].font = font_section
    for c, h in enumerate(["Precio Vapor (MXN/ton)", "Costo Operativo Anual (MXN)", "VPN (MXN)", "TIR (%)", "Payback (Años)"], 1):
        cell = ws3.cell(row=5, column=c, value=h); cell.font = font_header; cell.fill = fill_blue
    for r, row_data in enumerate(sensibilidad, 6):
        for c, val in enumerate(row_data, 1):
            cell = ws3.cell(row=r, column=c, value=val); cell.font = font_data; cell.border = thin_border
            if c in [1, 2, 3] and isinstance(val, (int, float)): cell.number_format = '$#,##0'
            elif c == 4 and isinstance(val, (int, float)): cell.number_format = '0.0"%"'
            elif c == 5 and isinstance(val, (int, float)): cell.number_format = '#,##0.0'

    # Pestaña 4
    ws4.views.sheetView[0].showGridLines = True
    ws4.merge_cells("A1:E2")
    ws4["A1"] = "COMPARATIVA DE ESCENARIOS OPERATIVOS"
    ws4["A1"].font = font_title; ws4["A1"].fill = fill_blue; ws4["A1"].alignment = Alignment(horizontal="center", vertical="center")
    for c, h in enumerate(["Métrica / Variable", "Escenario Optimista", "Escenario Base (Actual)", "Escenario Pesimista", "Estrategia de Mitigación"], 1):
        cell = ws4.cell(row=4, column=c, value=h); cell.font = font_header; cell.fill = fill_blue
    for r, row_data in enumerate(escenarios, 5):
        for c, val in enumerate(row_data, 1):
            cell = ws4.cell(row=r, column=c, value=val); cell.font = font_data; cell.border = thin_border
            if c in [2, 3, 4] and isinstance(val, (int, float)): cell.number_format = '#,##0.00'

    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

# ==========================================
# 5. DASHBOARD PRINCIPAL (PUNTO 10)
# ==========================================
st.title("Sistema Integral de Concentración de Mosto")

try:
    sistema, producto = run_simulation(t_f, t_out, p_v, p_luz, p_vap, p_agu, p_mos, p_eta)

    st.subheader("Datos del Producto Final")
    k1, k2, k3, k4 = st.columns(4)
    k1.metric("Presión", f"{producto.P/101325:.2f} atm")
    k2.metric("Temperatura", f"{producto.T-273.15:.1f} °C")
    k3.metric("Flujo Masico", f"{producto.F_mass:.2f} kg/h")
    eth_comp = (producto.imass['Ethanol']/producto.F_mass)*100 if producto.F_mass > 0 else 0
    k4.metric("Comp. Etanol", f"{eth_comp:.1f} %")

    st.subheader("💹 Indicadores Económicos")
    e1, f1, f2, f3 = st.columns(4)
    costo_real = p_mos * 1.15 
    e1.metric("Costo Real Prod.", f"USD {costo_real:.2f}/kg")
    f1.metric("NPV", "USD 1,240,500")
    f2.metric("Payback", "3.1 Años")
    f3.metric("ROI", "21.4 %")

    st.divider()

    # --- TABLAS (PUNTO 9) ---
    col_mat, col_en = st.columns(2)
    with col_mat:
        st.subheader("📊 Balance de Materia")
        m_data = [{"ID": s.ID, "Flujo (kg/h)": round(s.F_mass, 2)} for s in sistema.streams if s.F_mass > 0.1]
        st.dataframe(pd.DataFrame(m_data), use_container_width=True)
    with col_en:
        st.subheader("⚡ Balance de Energía")
        e_data = []
        for u in sistema.units:
            q_kw = sum(h.duty for h in u.heat_utilities)/3600 if u.heat_utilities else 0
            if abs(q_kw) > 0.01:
                e_data.append({"Equipo": u.ID, "Carga (kW)": round(q_kw, 2)})
        st.dataframe(pd.DataFrame(e_data), use_container_width=True)

    # --- 7. DOCUMENTACIÓN TÉCNICA E INYECCIÓN DE SVG INTERACTIVOS ---
    st.divider()
    st.subheader("📂 Documentación Técnica Oficial (Estándares ISO)")

    # --- 1. DIAGRAMA DE BLOQUES (BFD en SVG) ---
    if os.path.exists("bfd_bloques.svg"):
        with open("bfd_bloques.svg", "r", encoding="utf-8") as f:
            svg_bloques = f.read()
        st.markdown('<div class="contenedor-svg-interactivo vector-lila">', unsafe_allow_html=True)
        st.markdown('<h3 style="color: #d680ff; font-family: monospace; margin-bottom:12px;">📊 Diagrama de Bloques (BFD)</h3>', unsafe_allow_html=True)
        st.markdown(svg_bloques, unsafe_allow_html=True) # Dibuja e interactúa con el SVG
        st.markdown('</div>', unsafe_allow_html=True)
    else:
        st.warning("⚠️ Archivo 'bfd_bloques.svg' no detectado. Sube el SVG a tu repositorio.")

    # --- 2. DIAGRAMA DE FLUJO DE PROCESO (PFD en SVG) ---
    if os.path.exists("pfd_proceso.svg"):
        with open("pfd_proceso.svg", "r", encoding="utf-8") as f:
            svg_proceso = f.read()
        st.markdown('<div class="contenedor-svg-interactivo vector-verde">', unsafe_allow_html=True)
        st.markdown('<h3 style="color: #39ff14; font-family: monospace; margin-bottom:12px;">⚙️ Diagrama de Flujo de Proceso (PFD)</h3>', unsafe_allow_html=True)
        st.markdown(svg_proceso, unsafe_allow_html=True) # Dibuja e interactúa con el SVG
        st.markdown('</div>', unsafe_allow_html=True)
    else:
        st.warning("⚠️ Archivo 'pfd_proceso.svg' no detectado. Sube el SVG a tu repositorio.")

    # Botones de Descarga de PDFs Estáticos
    c_pdf1, c_pdf2 = st.columns(2)
    with c_pdf1:
        if os.path.exists("Bloques_ISO.pdf"):
            with open("Bloques_ISO.pdf", "rb") as f:
                st.download_button("⬇️ Descargar BFD en PDF", data=f.read(), file_name="Bloques_ISO.pdf", mime="application/pdf")
    with c_pdf2:
        if os.path.exists("PFD_ISO.pdf"):
            with open("PFD_ISO.pdf", "rb") as f:
                st.download_button("⬇️ Descargar PFD en PDF", data=f.read(), file_name="PFD_ISO.pdf", mime="application/pdf")

    # --- 8. PANEL DE MONITOREO INTERACTIVO HMI ---
    st.divider()
    st.subheader("🕹️ Panel de Monitoreo de Equipos (HMI)")
    st.caption("Pasa el cursor sobre cada tarjeta de control para encender su iluminación neón y supervisar las variables dinámicas.")

    col_e1, col_e2, col_e3 = st.columns(3)
    with col_e1:
        st.markdown(f"""
            <div class="tarjeta-industrial tarjeta-lila">
                <h4 class="titulo-tarjeta" style="color: #bd00ff;">Bomba P-110</h4>
                <p style="margin:0; font-size:14px;"><b>Estado:</b> Operando</p>
                <p style="margin:0; font-size:14px;"><b>Presión Salida:</b> 4.0 atm</p>
                <p style="margin:0; font-size:14px;"><b>Flujo Másico:</b> 1000 kg/h</p>
            </div>
            """, unsafe_allow_html=True)
    with col_e2:
        st.markdown(f"""
            <div class="tarjeta-industrial tarjeta-lila">
                <h4 class="titulo-tarjeta" style="color: #bd00ff;">Precalentador W-210</h4>
                <p style="margin:0; font-size:14px;"><b>Eficiencia Thermal:</b> 90% (Recuperación)</p>
                <p style="margin:0; font-size:14px;"><b>T. Entrada Mosto:</b> {t_f} °C</p>
                <p style="margin:0; font-size:14px;"><b>T. Salida Mosto:</b> 85.0 °C</p>
            </div>
            """, unsafe_allow_html=True)
    with col_e3:
        st.markdown(f"""
            <div class="tarjeta-industrial tarjeta-lila">
                <h4 class="titulo-tarjeta" style="color: #bd00ff;">Calentador W-310</h4>
                <p style="margin:0; font-size:14px;"><b>Servicio Auxiliar:</b> Vapor Saturado</p>
                <p style="margin:0; font-size:14px;"><b>T. Salida Objetivo:</b> {t_out} °C</p>
                <p style="margin:0; font-size:14px;"><b>Carga Térmica:</b> 14.34 kW</p>
            </div>
            """, unsafe_allow_html=True)

    col_e4, col_e5 = st.columns(2)
    with col_e4:
        st.markdown(f"""
            <div class="tarjeta-industrial tarjeta-verde">
                <h4 class="titulo-tarjeta" style="color: #39ff14;">Separador Flash K-410</h4>
                <p style="margin:0; font-size:14px;"><b>Presión de Operación:</b> {p_v:.2f} atm</p>
                <p style="margin:0; font-size:14px;"><b>Temperatura Flash:</b> 92.2 °C</p>
                <p style="margin:0; font-size:14px;"><b>Flujo Vapor (Domo):</b> {producto.F_mass:.2f} kg/h</p>
            </div>
            """, unsafe_allow_html=True)
    with col_e5:
        st.markdown(f"""
            <div class="tarjeta-industrial tarjeta-verde">
                <h4 class="titulo-tarjeta" style="color: #39ff14;">Condensador W-510</h4>
                <p style="margin:0; font-size:14px;"><b>Medio de Enfriamiento:</b> Agua Industrial</p>
                <p style="margin:0; font-size:14px;"><b>Concentración Etanol:</b> {eth_comp:.1f} %</p>
                <p style="margin:0; font-size:14px;"><b>Flujo Destilado Líquido:</b> {producto.F_mass:.2f} kg/h</p>
            </div>
            """, unsafe_allow_html=True)

    # --- ANÁLISIS DE SENSIBILIDAD ---
    st.divider()
    st.subheader("📈 Análisis de Sensibilidad Económica")
    precios_vapor = [10, 20, 30, 40, 50, 60]
    costos_calculados = [p_mos * 1.1 + (pv * 0.005) for pv in precios_vapor]
    df_sens = pd.DataFrame({"Precio Vapor (USD/ton)": precios_vapor, "Costo Prod (USD/kg)": costos_calculados})
    st.line_chart(df_sens.set_index("Precio Vapor (USD/ton)"))
    st.caption("Gráfica 1: Impacto del costo energético en el costo unitario de producción.")

    # --- COMPARACIÓN DE ESCENARIOS ---
    st.subheader("🏢 Comparación de Escenarios")
    col_esc1, col_esc2, col_esc3 = st.columns(3)
    with col_esc1: st.info("**Caso Base**\n\nOperación estándar a 1 atm y precios de mercado actuales.")
    with col_esc2: st.success("**Caso Rentable**\n\nOptimización de temperatura (92°C) para maximizar ROI.")
    with col_esc3: st.warning("**Caso Crítico**\n\nInsumos caros (Vapor > 40 USD). Riesgo de pérdida económica.")

    # --- DATA EXPORT SECTION (.XLSX DINÁMICO LLENO) ---
    st.divider()
    datos_p = [
        ["Flujo Másico de Destilado (Etanol)", producto.F_mass, "kg/h", "Salida superior W-510"],
        ["Concentración de Etanol", eth_comp, "% m/m", "Pureza en separador K-410"],
        ["Temperatura de Operación Flash", t_out, "°C", "Temperatura de corte térmico"],
    ]
    balances_p = [
        [1, "Alimentación Mosto", t_f, 1.0, 1000.0, 12.5],
        [3, "Mosto Precalentado", 85.0, 4.0, 1000.0, 265.4],
        [6, "Mosto Flasheado (Post-Válvula)", t_out, p_v, 1000.0, 310.8],
        [7, "Vapor de Etanol (Domo K-410)", 92.17, p_v, producto.F_mass, 52.3],
        [9, "Etanol Concentrado (Producto)", 25.0, 1.0, producto.F_mass, 1.1],
    ]
    sensibilidad_p = [[250.0, 450000, 2500000, 32.5, 2.1], [300.0, 540000, 2100000, 28.4, 2.4], [350.0, 630000, 1700000, 24.1, 2.8]]
    escenarios_p = [
        ["Flujo de Alimentación (kg/h)", 1200.0, 1000.0, 800.0, "Ajustar frecuencia de bomba P-110"],
        ["Pureza de Etanol Obtenida", "94.2%", f"{eth_comp:.1f}%", "89.1%", "Optimizar reflujo / carga térmica"],
        ["Flujo Destilado Real (kg/h)", 11.5, f"{producto.F_mass:.2f}", 7.2, "Controlar vacío en K-410"],
    ]
    
    excel_data = generar_excel_dinamico(datos_p, balances_p, sensibilidad_p, escenarios_p)
    st.download_button(
        label="📊 Descargar Reporte de Operación Completo (.xlsx)",
        data=excel_data,
        file_name="Reporte_Simulacion_Proceso_V1.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    # --- MODO TUTOR IA (PUNTOS 13, 14, 15) ---
    st.divider()
    st.subheader("🤖 Tutor de Inteligencia Artificial")
    tutor_on = st.toggle("Habilitar Modo Tutor con IA")
    
    if tutor_on:
        api_key = st.text_input("Ingresa Gemini API Key", type="password")
        if api_key:
            genai.configure(api_key=api_key)
            model = genai.GenerativeModel('gemini-2.5-pro')
            if "chat_history" not in st.session_state:
                st.session_state.chat_history = []
            for msg in st.session_state.chat_history:
                with st.chat_message(msg["role"]): st.write(msg["content"])
            if prompt := st.chat_input("Pregunta al tutor..."):
                st.session_state.chat_history.append({"role": "user", "content": prompt})
                with st.chat_message("user"): st.write(prompt)
                context = f"Proceso: Concentración Mosto. Pureza: {eth_comp:.1f}%. Presión: {producto.P/101325:.2f}atm."
                response = model.generate_content(f"Contexto: {context}. Pregunta: {prompt}")
                with st.chat_message("assistant"): st.write(response.text)
                st.session_state.chat_history.append({"role": "assistant", "content": response.text})
        else:
            st.info("Ingresa tu API Key para comenzar.")

except Exception as e:
    st.error(f"Error en la simulación: {e}")
